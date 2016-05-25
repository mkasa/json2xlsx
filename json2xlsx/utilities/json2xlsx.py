#!/usr/bin/env python2.7
# Requires pyparsing (ver 1.x; ver 2.x does not work), and openpyxl.
# Also requires csv and json.

import os, sys
from pyparsing import *
import argparse, subprocess
import openpyxl, json, csv

debugging = False

def upk(stuff):
    if len(stuff) <= 0:
        return None
    return stuff[0]

def merge_dict(d1, d2):
    if d2 == None: return d1
    if type(d2) == dict:
        for k, v in d2.iteritems():
            d1[k] = v
    elif type(d2) == type([]):
        for d in d2:
            for k, v in d.iteritems():
                d1[k] = v
    else:
        print "Assertion Error: Missing .asList()?", type(d2), type([])
        sys.exit(99)
    return d1

LPAREN = Suppress("(")
RPAREN = Suppress(")")
LBRACE = Suppress("{")
RBRACE = Suppress("}")
COMMA = Suppress(",")
DELIMITER = Suppress(";") | Suppress(",")
BareString = Word(alphas + "._", alphanums + "._")
NumberString = Optional("-") + Word("0123456789")
NumberString.setParseAction( lambda s,l,t: t[0] + t[1] if 1 < len(t) else t[0] )
QuoteString = QuotedString(quoteChar="\"", escChar="\\")
QuoteString.setParseAction( lambda s,l,t: t[0].replace("\\n", "\n") )
LiteralString = QuoteString | BareString
ColorAddition = Suppress(Keyword("color")) - LiteralString
ColorAddition.setParseAction( lambda s,l,t: {"color": t[0].upper()} )
BorderAddition = Suppress(Keyword("border")) - LiteralString
BorderAddition.setParseAction( lambda s,l,t: {"border": t[0]} )
AlignmentAddition = Suppress(Keyword("align")) - LiteralString
AlignmentAddition.setParseAction( lambda s,l,t: {"align": t[0]} )
HeadingAlignmentAddition = Suppress(Keyword("halign")) - LiteralString
HeadingAlignmentAddition.setParseAction( lambda s,l,t: {"halign": t[0]} )
NumberStyleAddition = Suppress(Keyword("number")) - LiteralString
NumberStyleAddition.setParseAction( lambda s,l,t: {"number": t[0]} )
ColumnWidthAddition = Suppress(Keyword("width")) - NumberString
ColumnWidthAddition.setParseAction( lambda s,l,t: {"column_width": t[0]} )
AttributeProperty = ColorAddition | AlignmentAddition | HeadingAlignmentAddition | NumberStyleAddition | ColumnWidthAddition
AttributeStatement = LiteralString - Group(Optional(Suppress("as") - LiteralString)) - Group(ZeroOrMore(AttributeProperty))
AttributeStatement.setParseAction( lambda s,l,t: \
        merge_dict({"type": "attr", "select": t[0], "caption": upk(t[1])}, t[2].asList()) )
GroupBlock = Forward()
GroupProperty = ColorAddition | BorderAddition
GroupStatement = Suppress(Keyword("group")) - LiteralString - \
    Group(Optional(LiteralString)) - Group(ZeroOrMore(GroupProperty)) - Group(GroupBlock)
GroupStatement.setParseAction( lambda s,l,t: \
        merge_dict({"type": "group", "caption": t[0], "select": upk(t[1]), "content": t[3].asList()}, t[2].asList()) )

TableStatement = (GroupStatement + Optional(DELIMITER)) | (AttributeStatement + DELIMITER)
TableStatements = ZeroOrMore(TableStatement)
TableBlock = LBRACE - TableStatements - RBRACE
GroupBlock << TableBlock
TableProperty = BorderAddition
TableDeclaration = Suppress(Keyword("table")) - Group(Optional(LiteralString)) - Group(Optional(ColorAddition)) - Group(ZeroOrMore(TableProperty)) - TableBlock
TableDeclaration.setParseAction( lambda s,l,t: \
        merge_dict({"type": "table", "content": t[3:], "caption": upk(t[0]), "color": upk(t[1])}, t[2].asList()) )
LoadFromFileStatement = Suppress(Keyword("load")) - LiteralString - Group(Optional(Suppress("as") | LiteralString)) - Group(Optional("linebyline")) - DELIMITER
LoadFromFileStatement.setParseAction( lambda s,l,t: {"type": "load", "filename": t[0], "caption": upk(t[1]), "line_by_line": upk(t[2])} )
LoadCSVFromFileStatement = Suppress(Keyword("loadcsv")) - LiteralString - Group(Optional(Suppress("as") |\
        LiteralString)) - Group(Optional(delimitedList(NumberString))) - Group(Optional("withheader")) - DELIMITER
LoadCSVFromFileStatement.setParseAction( lambda s,l,t: {"type": "loadcsv", "filename": t[0],\
        "caption": upk(t[1]), "column_order": t[2], "withheader": upk(t[3])} )
WriteToFileStatement = Suppress(Keyword("write") | Keyword("save")) - LiteralString - Group(Optional(Keyword("open"))) - DELIMITER
WriteToFileStatement.setParseAction( lambda s,l,t: {"type": "save", "filename": t[0], "open": upk(t[1])} )
ShowHeaderStatement = Suppress(Keyword("header")) - DELIMITER
ShowHeaderStatement.setParseAction( lambda s,l,t: {"type": "header"} )
VSkipStatement = Suppress(Keyword("vskip")) - NumberString - DELIMITER
VSkipStatement.setParseAction( lambda s,l,t: {"type": "vskip", "amount": int(t[0])} )
NewSheetStatement = Suppress(Keyword("newsheet")) - Group(Optional(NumberString)) - DELIMITER
NewSheetStatement.setParseAction( lambda s,l,t: {"type": "newsheet", "position": None if len(t[0]) < 1 else int(t[0])} )
NameSheetStatement = Suppress(Keyword("namesheet")) - LiteralString - DELIMITER
NameSheetStatement.setParseAction( lambda s,l,t: {"type": "namesheet", "name": t[0]} )
WriteLegendStatement = Suppress(Keyword("legend") | Keyword("writelegend")) - NumberString - Suppress(COMMA) - NumberString - LiteralString - DELIMITER
WriteLegendStatement.setParseAction( lambda s,l,t: {"type": "writelegend", "yoffset": int(t[0]), "xoffset": int(t[1]), "value": t[2]} )

Operation = TableDeclaration | LoadFromFileStatement | LoadCSVFromFileStatement | WriteToFileStatement\
        | ShowHeaderStatement | VSkipStatement | NewSheetStatement | NameSheetStatement \
        | WriteLegendStatement
TableScript = ZeroOrMore(Operation) + StringEnd()
table_line_comment = "#" + restOfLine
table_cpp_line_comment = "//" + restOfLine
TableScript.ignore(table_line_comment)
TableScript.ignore(table_cpp_line_comment)

class SizeRenderingError(Exception):
    pass

def size_render(table_scr_tree):
    y_size = 0; x_size = 0
    for node in table_scr_tree:
        if debugging: print "Size Proc: ", node
        node_type = node['type']
        if node_type == 'attr':
            x_size += 1
            y_size = max(y_size, 1)
        elif node_type == 'table':
            (child_y, child_x) = size_render(node['content'])
            x_size += child_x
            if node['caption'] == None:
                y_size = max(y_size, child_y)
            else:
                y_size = max(y_size, child_y + 1)
        elif node_type == 'group':
            (child_y, child_x) = size_render(node['content'])
            x_size += child_x
            y_size = max(y_size, child_y + 1)
        else:
            raise SizeRenderingError()
    return y_size, x_size

class RenderingError(Exception):
    pass

def set_cell_color_if_needed(cell, color_string):
    if color_string == None or color_string == "": return
    try:
        cell.style.font.color.index = openpyxl.style.Color.__dict__[color_string]
    except:
        raise RenderingError("Unknown color string '%s'" % color_string)

def set_cell_value_and_wrap_if_needed(cell, content):
    cell.value = content
    if "\n" in content: cell.style.alignment.wrap_text = True

def set_cell_align_if_needed(cell, align_string):
    if align_string == None or align_string == "": return
    if align_string == "right":
        cell.style.alignment.horizontal = openpyxl.style.Alignment.HORIZONTAL_RIGHT
    elif align_string == "left":
        cell.style.alignment.horizontal = openpyxl.style.Alignment.HORIZONTAL_LEFT
    elif align_string == "center":
        cell.style.alignment.horizontal = openpyxl.style.Alignment.HORIZONTAL_CENTER
    else:
        raise RenderingError("Unknown align string '%s'" % align_string)

def set_cell_number_style_if_needed(cell, number_style):
    if number_style == None: return
    if number_style == ",":
        cell.style.number_format.format_code = "#,##"
        return
    cell.style.number_format.format_code = number_style

def set_column_width_if_needed(worksheet, column, width_style):
    if width_style == None: return
    worksheet.column_dimensions[openpyxl.cell.get_column_letter(column + 1)].width = width_style

def set_range_border_if_needed(worksheet, yrange, xrange, border_style):
    if border_style == None or border_style == "": return
    cell_style = None
    if border_style == "thinbottom":   cell_style = openpyxl.style.Border.BORDER_THIN
    if border_style == "thickbottom":  cell_style = openpyxl.style.Border.BORDER_THICK
    if border_style == "doublebottom": cell_style = openpyxl.style.Border.BORDER_DOUBLE
    if border_style == None:
        raise RenderingError("Unknown border style '%s'" % border_style)
    for x in range(xrange[0], xrange[1] + 1):
        worksheet.cell(row = yrange[1] + 1, column = x + 1).style.borders.bottom.border_style = border_style

def render(workbook, cursor, y_range, x_range, render_state, tree):
    for node in tree:
        if debugging: print "Render ", node
        if node == None:
            raise RenderingError("Table is not yet defined")
        current_sheet = render_state['current_sheet']
        if debugging: print "Process ", node
        node_type = node['type']
        if node_type == 'attr':
            cell = current_sheet.cell(row = cursor[0] + y_range - 1 + 1, column = cursor[1] + 1)
            caption = node.get('caption')
            if caption == None: caption = node['select']
            set_cell_value_and_wrap_if_needed(cell, caption)
            set_cell_align_if_needed(cell, node.get('halign'))
            set_column_width_if_needed(current_sheet, cursor[1], node.get('column_width'))
            render_state['column_to_attr'][cursor[1]] = {"align": node.get('align'), "number": node.get('number'), "color": node.get('color')}
            cursor[1] += 1
        elif node_type == 'table':
            cell = current_sheet.cell(row = cursor[0] + 1, column = cursor[1] + 1)
            caption = node['caption']
            new_cursor = [cursor[0], cursor[1]]
            new_y_range = y_range
            if caption != None:
                set_cell_value_and_wrap_if_needed(cell, caption)
                set_cell_color_if_needed(cell, node.get('color'))
                current_sheet.merge_cells(start_row = cursor[0] + 1, start_column = cursor[1] + 1, end_row = cursor[0] + 1, end_column = cursor[1] + x_range - 1 + 1)
                set_range_border_if_needed(current_sheet, [cursor[0], cursor[0] + y_range - 1], [cursor[1], cursor[1] + x_range - 1], node.get('border'))
                new_cursor[0] += 1
                new_y_range -= 1
            children = node['content']
            render(workbook, new_cursor, new_y_range, x_range, render_state, children)
            render_state['table.left']  = cursor[1]
            render_state['table.right'] = cursor[1] + x_range - 1
            render_state['table.top']   = cursor[0]
            render_state['table.header.top']    = cursor[0]
            render_state['table.header.bottom'] = cursor[0] + y_range - 1
            render_state['table.data.top']      = cursor[0] + y_range
            cursor[0] += y_range
        elif node_type == 'group':
            children = node['content']
            (child_y, child_x) = size_render(children)
            if debugging: print "Child size (%d x %d)" % (child_y, child_x)
            cell = current_sheet.cell(row = cursor[0] + 1, column = cursor[1] + 1)
            caption = node['caption']
            set_cell_value_and_wrap_if_needed(cell, caption)
            set_cell_color_if_needed(cell, node.get('color'))
            current_sheet.merge_cells(start_row = cursor[0] + 1, start_column = cursor[1] + 1, end_row = cursor[0] + 1, end_column = cursor[1] + child_x - 1 + 1)
            set_range_border_if_needed(current_sheet, [cursor[0], cursor[0] + y_range - 1], [cursor[1], cursor[1] + child_x - 1], node.get('border'))
            if debugging: print "merge (%d, %d) - (%d, %d)" % (cursor[0], cursor[1], cursor[0], cursor[1] + child_x - 1)
            new_cursor = [cursor[0] + 1, cursor[1]]
            render(workbook, new_cursor, y_range - 1, child_x, render_state, children)
            cursor[1] += child_x
        else:
            raise RenderingError()

class RenderingDataError(Exception):
    pass

def select_json(json_object, select_stmt):
    if select_stmt == None: return json_object
    if select_stmt == "": return ""
    select_series = select_stmt.split(".")
    try:
        for st in select_series:
            if st != "":
                json_object = json_object[st]
    except:
        raise RenderingDataError("No such attr '%s'" % select_stmt)
    return json_object

def render_csv_data(workbook, cursor, render_state, column_order, csv):
    if debugging: print "Render CSV Data ", ','.join(csv)
    current_sheet = render_state['current_sheet']
    if debugging: print "Column Order: ", column_order, len(column_order)
    if len(column_order) == 0:
        column_order = [x for x in range(0, len(csv))]
    else:
        column_order = [int(x) for x in column_order]
    for index in range(len(column_order)):
        if debugging: print "Process ", value
        cell = current_sheet.cell(row = cursor[0] + 1, column = cursor[1] + 1)
        csv_index = column_order[index]
        if 0 <= csv_index:
            cell.value = csv[csv_index]
        else:
            cell.value = ""
        try:
            cell_style = render_state['column_to_attr'][cursor[1]]
        except:
            raise RenderingDataError("No such column (%d) in a header. Header has not been printed yet." % cursor[1])
        set_cell_color_if_needed(cell, cell_style['color'])
        set_cell_align_if_needed(cell, cell_style['align'])
        set_cell_number_style_if_needed(cell, cell_style['number'])
        cursor[1] += 1
    cursor[0] += 1
    cursor[1] = 0

def render_data(workbook, cursor, render_state, tree):
    if debugging: print "Render Data ", tree
    current_sheet = render_state['current_sheet']
    for node in tree:
        if debugging: print "Process ", node
        cell = current_sheet.cell(row = cursor[0] + 1, column = cursor[1] + 1)
        node_type = node['type']
        if node_type == 'attr':
            json_node = render_state['json_object']
            cell.value = select_json(json_node, node['select'])
            try:
                cell_style = render_state['column_to_attr'][cursor[1]]
            except:
                raise RenderingDataError("No such column (%d) in a header. Header has not been printed yet." % cursor[1])
            set_cell_color_if_needed(cell, cell_style['color'])
            set_cell_align_if_needed(cell, cell_style['align'])
            set_cell_number_style_if_needed(cell, cell_style['number'])
            cursor[1] += 1
        elif node_type == 'table':
            render_data(workbook, cursor, render_state, node['content'])
            cursor[0] += 1
            cursor[1] = 0
        elif node_type == 'group':
            save_json_obj = render_state['json_object']
            render_state['json_object'] = select_json(save_json_obj, node.get('select'))
            render_data(workbook, cursor, render_state, node['content'])
            render_state['json_object'] = save_json_obj
        else:
            raise RenderingDataError()

def parse_table_script(table_script_file_name):
    try:
        if table_script_file_name == '-':
            stdin_string = sys.stdin.read()
            table_scr_tree = TableScript.parseString(stdin_string)
        else:
            table_scr_tree = TableScript.parseFile(table_script_file_name)
        # print table_scr_tree
        return table_scr_tree
    except ParseException, e:
        print "Table script error:"
        print e.line
        print " " * (e. column - 1) + "^"
        print e
        sys.exit(1)
    except ParseSyntaxException, e:
        print "Table script syntax error:"
        print e.line
        print " " * (e. column - 1) + "^"
        print e
        sys.exit(2)

def main_real():
    parser = argparse.ArgumentParser(description="JSON to Excel table")
    parser.add_argument('tablescript', help='table script')
    parser.add_argument('-e', help='execute a rendering command')
    parser.add_argument('-r', help='rendering script file')
    parser.add_argument('-o', help='output file')
    parser.add_argument('--open', action='store_true', help='open the generated file immediately')
    parser.add_argument('-l', action='store_true', help='use one-line-one-JSON-object file as input')
    parser.add_argument('-j', action='append', help='input json')
    parser.add_argument('-n', action='append', help='name for inputs (This option must be repeated as many times as -j')
    args = parser.parse_args()

    table_scr_tree = parse_table_script(args.tablescript)

    workbook = openpyxl.Workbook()
    cursor = [0, 0]
    has_anything_output = [False]
    render_state = {"current_sheet": workbook.get_active_sheet(),\
            "column_to_attr": {}, "header_needed": False,\
            "current_table": None}

    def write_the_book_to_file(file_name):
        if debugging: print "Saving to '%s'" % file_name
        workbook.save(filename = file_name)
        has_anything_output[0] = True

    def load_from_csv_file_and_render(file_name, column_order, caption, has_header):
        try:
            with open(file_name, 'r') as csvfile:
                reader = csv.reader(csvfile)
                if has_header:
                    dummy_line = reader.readline()
                for row in reader:
                    render_csv_data(workbook, cursor, render_state, column_order, row)
        except IOError as e:
            print >>sys.stderr, "ERROR: could not open '%s'. %s" % (file_name, e.strerror)
            sys.exit(4)

    def load_from_json_file_and_render(file_name, caption, one_json_obj_per_line):
        if not one_json_obj_per_line:
            try:
                json_obj = json.load(open(file_name, "r"))
            except IOError as a:
                print >>sys.stderr, "ERROR: could not open '%s'" % file_name
                sys.exit(4)
            except ValueError as e:
                print >>sys.stderr, "ERROR: JSON Parsing", e
                print >>sys.stderr, "       This error may occur when the input file contain multiple JSON objects."
                print >>sys.stderr, "       When you want to include multiple JSON objects in a single file,"
                print >>sys.stderr, "       Give '-l' to json2xlsx, and place a single JSON object in each line in an input file."
                print >>sys.stderr, "       That above restriction may sound too strict, but removing this restriction requires"
                print >>sys.stderr, "       significant reengineering of the standard JSON module of Python."
        else:
            try:
                file_obj = open(file_name, "r")
                line_number = 0
            except:
                print >>sys.stderr, "ERROR: could not open '%s'" % file_name
                sys.exit(6)
        while True:
            try:
                if one_json_obj_per_line:
                    line_number += 1
                    try:
                        line_str = file_obj.readline().strip()
                    except:
                        print >>sys.stderr, "ERROR: read error at line ", line_number
                        sys.exit(7)
                    if line_str == "": return
                    json_obj = json.loads(line_str)
            except:
                print >>sys.stderr, "ERROR: JSON parsing error at line ", line_number
                sys.exit(5)
            if type(json_obj) == type([]):
                for i, child in enumerate(json_obj):
                    if(child) is dict:
                        child['file_name'] = file_name
                        child['file_caption'] = caption
                        child['file_index'] = i
                    render_state['json_object'] = child
                    render_data(workbook, cursor, render_state, [render_state['current_table']])
            else:
                if type(json_obj) is dict:
                    json_obj['file_name'] = file_name
                    json_obj['file_caption'] = caption
                render_state['json_object'] = json_obj
                render_data(workbook, cursor, render_state, [render_state['current_table']])
            if not one_json_obj_per_line: break

    def interpret_render_scr_tree(node):
        node_type = node['type']
        if debugging: print "Node: ", node
        if node_type == 'table':
            render_state['header_needed'] = True
            render_state['current_table'] = node
        elif node_type == 'header' or node_type == 'load' or node_type == 'loadcsv':
            if node_type == 'header' or ((node_type == 'load' or node_type == 'loadcsv') and render_state['header_needed']):
                (y_size, x_size) = size_render([render_state['current_table']])
                if debugging: print "Size (%d x %d)" % (y_size, x_size)
                render(workbook, cursor, y_size, x_size, render_state, [render_state['current_table']])
                render_state['header_needed'] = False
            if node_type == 'load':
                load_from_json_file_and_render(node['filename'], node['caption'], node['linebyline'] != None)
            elif node_type == 'loadcsv':
                load_from_csv_file_and_render(node['filename'], node['column_order'], node['caption'], node['has_header'])
        elif node_type == 'save':
            write_the_book_to_file(node['filename'])
            if node['open'] != None: subprocess.call(["open", node['filename']])
        elif node_type == 'vskip':
            cursor[0] += node['amount']
        elif node_type == 'newsheet':
            if node['position'] == None:
                render_state['current_sheet'] = workbook.create_sheet()
            else:
                render_state['current_sheet'] = workbook.create_sheet(node['position'])
            cursor[0] = 0; cursor[1] = 0
        elif node_type == 'namesheet':
            render_state['current_sheet'].title = node['name']
        elif node_type == 'writelegend':
            render_state['current_sheet'].cell(row = render_state['table.top'] + node['yoffset'], column = render_state['table.right'] + 1 + node['xoffset']).value = node['value']
        else:
            raise RenderingError("Unknown rendering command '%s' (this is a bug; please report it to the author)" % node_type)

    for node in table_scr_tree:
        interpret_render_scr_tree(node)

    if args.j != None:
        if args.n != None and len(args.j) != len(args.n):
            print >>sys.stderr, "ERROR: -n must be given as exactly the same time as -j"
            sys.exit(1)
        for json_file, caption in zip(args.j, args.n if args.n != None else [None] * len(args.j)):
            if debugging: print "Render ", json_file
            if render_state['header_needed']:
                (y_size, x_size) = size_render([render_state['current_table']])
                render(workbook, cursor, y_size, x_size, render_state, [render_state['current_table']])
                render_state['header_needed'] = False
            load_from_json_file_and_render(json_file, json_file if caption == None else caption, args.l)

    if args.o != None or args.j != None:
        if args.o == None and args.j != None:
            args.o = args.j[0] + ".xlsx"
        write_the_book_to_file(args.o)
        if args.open:
            subprocess.call(["open", args.o])

    if not has_anything_output[0]:
        print "Please give -o option to specify the name of the output file, or put 'write' command in a rendering script"
        print "e.g.1) give '-o output.xlsx' to command line option"
        print "e.g.2) add 'write \"output.xlsx\"' (without single quotations) to an input rendering script"
        sys.exit(9)

def main():
    try:
        main_real()
    except KeyboardInterrupt:
        print ""
        sys.exit()

if __name__ == "__main__":
    main()
